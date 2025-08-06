import logging
from aiogram.enums import ParseMode
from aiogram import *
from aiogram.types import *
from aiogram.client.bot import *
from aiogram.filters import *
from aiogram.utils.keyboard import *
from aiogram.types.input_media_photo import *
from aiogram.filters.callback_data import *
from aiogram.exceptions import *
import sqlite3
import asyncio
import datetime
from datetime import datetime, timedelta
from collections import defaultdict
import calendar
from decimal import Decimal, ROUND_HALF_UP
import time
import openpyxl
from openpyxl.utils import get_column_letter
from collections import defaultdict
import re
import os
from io import BytesIO

print('Jiji_TEST')

admins = [1042304148, 868006823, 1796733875, 890441921, 844665324] # 
TOKEN = "1424518106:AAEGqRChBBsrHNtMvom6eh1K4YhqZZ3DeYs"
"""# JIJI
channel = '-1001698914820' # -1796733875
link = 'https://t.me/c/1698914820/'"""
#test
channel = '-1001962474362'
link = 'https://t.me/c/1962474362/'

dp = Dispatcher()

menu = [[InlineKeyboardButton(text='Пополнение', callback_data='add')],
         [InlineKeyboardButton(text='Реализация', callback_data='delete')],
         [InlineKeyboardButton(text='Бухгалтерия', callback_data='sklad')],
         ]

show_menu = InlineKeyboardMarkup(inline_keyboard=menu)

stat = [[InlineKeyboardButton(text='Текущий месяц', callback_data='this')],
        [InlineKeyboardButton(text='Прошлый месяц', callback_data='last')],
        [InlineKeyboardButton(text='Общая статистика', callback_data='all')],
        [InlineKeyboardButton(text='Наличие', callback_data='buy')],
        [InlineKeyboardButton(text='Отменить продажу', callback_data='cancelLast')],
        [InlineKeyboardButton(text='Назад', callback_data='sklad'), InlineKeyboardButton(text='МЕНЮ', callback_data='menu')]
         ]

show_stat = InlineKeyboardMarkup(inline_keyboard=stat)

# create a logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# create a file handler
handler = logging.FileHandler('Jiji TEST.log', encoding='utf-8')
handler.setLevel(logging.INFO)

# create a logging format
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s', datefmt='%d-%m-%Y')
handler.setFormatter(formatter)

# add the handler to the logger
logger.addHandler(handler)

async def connect():
    global db, sql
    db = sqlite3.connect('Jiji TEST.db')
    sql = db.cursor()

db_keyboard = [
    [InlineKeyboardButton(text='Получить', callback_data='take')],
    #[InlineKeyboardButton(text='Отправить', callback_data='send')],
    [InlineKeyboardButton(text='Назад', callback_data='sklad'), InlineKeyboardButton(text='МЕНЮ', callback_data='menu')]
    ]
show_db_keyboard = InlineKeyboardMarkup(inline_keyboard=db_keyboard)

send = 0

@dp.callback_query(lambda query: query.data == 'database')
async def database(query):
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup=show_menu)
    await clear()
    await query.message.edit_text('База данных', reply_markup=show_db_keyboard)

@dp.callback_query(lambda query: query.data == 'take')
async def take_db(query):
    global send
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup=show_menu)
    await clear()
    db_file_path = 'Jiji TEST.db'
    try:
        await query.message.answer_document(FSInputFile(db_file_path))
        await query.message.answer('База данных', reply_markup=show_db_keyboard)
    except:
        await query.message.edit_text('Файл базы данных не найден.', reply_markup=show_menu)

'''@dp.callback_query(lambda query: query.data == 'send')
async def send_db(query):
    global send
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup=show_menu)
    send = 1
    await query.message.answer('Отправьте файл следующим сообщением.', reply_markup=show_db_keyboard)

@dp.message(F.content_type == ContentType.DOCUMENT)
async def save_db(message):
    global send
    if message.from_user.id not in admins:
        return await message.answer('Ты не мой братик!', reply_markup=show_menu)
    if send == 0: return
    db.close()
    db_file_path = 'Jiji TEST.db'
    if os.path.exists(db_file_path):
        new_file_name = f'Jiji_{datetime.now().strftime('%Y.%m.%d_%H-%M-%S')}.db'
        os.rename(db_file_path, new_file_name)

    file_id = message.document.file_id
    file_path = await bot.get_file(file_id)
    downloaded_file = await bot.download_file(file_path.file_path)
    
    with open(db_file_path, 'wb') as f:
        f.write(downloaded_file.read())
    
    await message.answer('Файл базы данных успешно обновлен.', reply_markup=show_menu)
    await clear()
'''
#-----------------------------------------------------------------------------------------------
@dp.callback_query(lambda query: query.data == 'stat')
async def statistics(query):
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup = show_menu)
    await query.message.edit_text(f'Статистика', reply_markup=show_stat)

async def write_sales_this(sales_data, total_sales, total_profit, type_sales, type_profit):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers = ['Время', 'Дата', 'Название', 'Тип', 'Характеристика', 'Цена', 'Профит', 'Приписка']
    sheet.append(headers)

    for sale in sales_data:
        num, product_name, product_type, product_x1, product_x2, product_v1, purchase_price, selling_price, note, sale_date, sale_time = sale
        formatted_sale_date = datetime.strptime(sale_date, '%Y.%m.%d').strftime('%m.%d')
        profit = selling_price - purchase_price
        row_data = [sale_time, formatted_sale_date, product_name, product_type, product_v1, selling_price, profit, note]
        sheet.append(row_data)

    sheet.cell(row=1, column=10, value='Всего реализовано:')
    sheet.cell(row=1, column=11, value=total_sales)
    sheet.cell(row=2, column=10, value='Выручка (прибыль):')
    sheet.cell(row=2, column=11, value=total_profit)

    row_offset = 3
    sheet.cell(row=row_offset - 2, column=13, value='По типам продуктов')
    sheet.cell(row=row_offset - 1, column=13, value='Тип')
    sheet.cell(row=row_offset - 1, column=14, value='Количество')
    sheet.cell(row=row_offset - 1, column=15, value='Выручка')

    for i, (product_type, sales_count) in enumerate(type_sales.items(), start=row_offset):
        sheet.cell(row=i, column=13, value=f'{product_type}:')
        sheet.cell(row=i, column=14, value=sales_count)
        sheet.cell(row=i, column=15, value=type_profit[product_type])

    column_widths = [8, 8, 25, 23, 30, 8, 8, 30, 10, 23, 11, 10, 23, 13, 10]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    excel_file_name = 'this_stats.xlsx'
    workbook.save(excel_file_name)

    return excel_file_name

@dp.callback_query(lambda query: query.data == 'this')
async def stat_this(query):
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup = show_menu)
    current_date = datetime.now()

    first_day_of_month = current_date.replace(day=1)
    last_day_of_month = first_day_of_month.replace(
        day=calendar.monthrange(current_date.year, current_date.month)[1]
    )

    first_day_of_month_str = first_day_of_month.strftime('%Y.%m.%d')
    last_day_of_month_str = last_day_of_month.strftime('%Y.%m.%d')

    sql_query = f'''SELECT * FROM Stat WHERE date BETWEEN '{first_day_of_month_str}' AND '{last_day_of_month_str}' '''
    sql.execute(sql_query)
    sales_data = sql.fetchall()

    total_sales = 0
    total_revenue = 0
    total_profit = 0

    if sales_data:
        response_message = f'Продажи за {current_date.strftime('%m.%y')}:\n'
        type_sales = defaultdict(int)
        type_revenue = defaultdict(float)
        type_profit = defaultdict(float)

        for sale in sales_data:
            num, product_name, product_type, product_x1, product_x2, product_v1, purchase_price, selling_price, note, sale_date, sale_time = sale
        
            total_sales += 1
            total_revenue += selling_price
            total_revenue = int(total_revenue) if total_revenue.is_integer() else round(total_revenue, 2)

            total_profit += (selling_price - purchase_price)
            total_profit = int(total_profit) if total_profit.is_integer() else round(total_profit, 2)

            type_sales[product_type] += 1

            type_revenue[product_type] += selling_price
            type_revenue[product_type] = int(type_revenue[product_type]) if type_revenue[product_type].is_integer() else round(type_revenue[product_type], 2)

            type_profit[product_type] += (selling_price - purchase_price)
            type_profit[product_type] = int(type_profit[product_type]) if type_profit[product_type].is_integer() else round(type_profit[product_type], 2)

        response_message += f'Всего реализовано: {total_sales}\n'
        response_message += f'Выручка (прибыль): {total_revenue}р ({total_profit}р)\n'

        response_message += '➖➖➖➖➖➖➖➖➖➖\n'
        for product_type, sales_count in type_sales.items():
            response_message += f'{product_type}\t: {sales_count}\t| {type_revenue[product_type]}р ({type_profit[product_type]}р)\n'

        excel_file = await write_sales_this(sales_data, total_sales, total_profit, type_sales, type_profit)

    else:
        response_message = 'Нет продаж за текущий месяц.'

    await query.message.answer_document(FSInputFile(excel_file), caption=response_message, parse_mode=ParseMode.MARKDOWN)
    await query.message.answer('Стартовое меню', reply_markup=show_menu)

async def write_sales_last(sales_data, total_sales, total_profit, type_sales, type_profit):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers = ['Время', 'Дата', 'Название', 'Тип', 'Характеристика', 'Цена', 'Профит', 'Приписка']
    sheet.append(headers)

    for sale in sales_data:
        num, product_name, product_type, product_x1, product_x2, product_v1, purchase_price, selling_price, note, sale_date, sale_time = sale
        formatted_sale_date = datetime.strptime(sale_date, '%Y.%m.%d').strftime('%m.%d')
        profit = selling_price - purchase_price
        row_data = [sale_time, formatted_sale_date, product_name, product_type, product_v1, selling_price, profit, note]
        sheet.append(row_data)

    sheet.cell(row=1, column=10, value='Всего продаж:')
    sheet.cell(row=1, column=11, value=total_sales)
    sheet.cell(row=2, column=10, value='Общая прибыль:')
    sheet.cell(row=2, column=11, value=total_profit)

    row_offset = 3
    sheet.cell(row=row_offset - 2, column=13, value='По типам продуктов')
    sheet.cell(row=row_offset - 1, column=13, value='Тип')
    sheet.cell(row=row_offset - 1, column=14, value='Количество')
    sheet.cell(row=row_offset - 1, column=15, value='Выручка')

    for i, (product_type, sales_count) in enumerate(type_sales.items(), start=row_offset):
        sheet.cell(row=i, column=13, value=f'{product_type}:')
        sheet.cell(row=i, column=14, value=sales_count)
        sheet.cell(row=i, column=15, value=type_profit[product_type])

    column_widths = [8, 8, 25, 23, 30, 8, 8, 30, 10, 23, 11, 10, 23, 13, 10, 30]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    excel_file_name = 'last_stats.xlsx'
    workbook.save(excel_file_name)

    return excel_file_name

@dp.callback_query(lambda query: query.data == 'last')
async def stat_last(query):
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup = show_menu)
    current_date = datetime.now()

    first_day_of_previous_month = (current_date.replace(day=1) - timedelta(days=1)).replace(day=1)
    last_day_of_previous_month = first_day_of_previous_month.replace(
        day=calendar.monthrange(first_day_of_previous_month.year, first_day_of_previous_month.month)[1]
    )

    first_day_of_month_str = first_day_of_previous_month.strftime('%Y.%m.%d')
    last_day_of_month_str = last_day_of_previous_month.strftime('%Y.%m.%d')

    sql_query = f'''SELECT * FROM Stat WHERE date BETWEEN '{first_day_of_month_str}' AND '{last_day_of_month_str}' '''
    sql.execute(sql_query)
    sales_data = sql.fetchall()

    total_sales = 0
    total_revenue = 0
    total_profit = 0

    if sales_data:
        response_message = f'Продажи за {first_day_of_previous_month.strftime('%m.%y')}:\n'
        type_sales = defaultdict(int)
        type_revenue = defaultdict(float)
        type_profit = defaultdict(float)

        for sale in sales_data:
            num, product_name, product_type, product_x1, product_x2, product_v1, purchase_price, selling_price, note, sale_date, sale_time = sale
        
            total_sales += 1
            total_revenue += selling_price
            total_revenue = int(total_revenue) if total_revenue.is_integer() else round(total_revenue, 2)

            total_profit += (selling_price - purchase_price)
            total_profit = int(total_profit) if total_profit.is_integer() else round(total_profit, 2)

            type_sales[product_type] += 1

            type_revenue[product_type] += selling_price
            type_revenue[product_type] = int(type_revenue[product_type]) if type_revenue[product_type].is_integer() else round(type_revenue[product_type], 2)

            type_profit[product_type] += (selling_price - purchase_price)
            type_profit[product_type] = int(type_profit[product_type]) if type_profit[product_type].is_integer() else round(type_profit[product_type], 2)

        response_message += f'Всего реализовано: {total_sales}\n'
        response_message += f'Выручка (прибыль): {total_revenue}р ({total_profit}р)\n'

        response_message += '➖➖➖➖➖➖➖➖➖➖\n'
        for product_type, sales_count in type_sales.items():
            response_message += f'{product_type}\t: {sales_count}\t| {type_revenue[product_type]}р ({type_profit[product_type]}р)\n'

        excel_file = await write_sales_last(sales_data, total_sales, total_profit, type_sales, type_profit)

    else:
        response_message = 'Нет продаж за предыдущий месяц.'

    await query.message.answer_document(FSInputFile(excel_file), caption=response_message, parse_mode=ParseMode.MARKDOWN)
    await query.message.answer('Что делаем?', reply_markup=show_menu)

async def write_sales_to_excel(sales_data, total_sales, total_profit, first_sale_date, type_sales, type_profit):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers = ['Время', 'Дата', 'Название', 'Тип', 'Характеристика', 'Цена', 'Профит', 'Приписка']
    sheet.append(headers)

    for sale in sales_data:
        num, product_name, product_type, product_x1, product_x2, product_v1, purchase_price, selling_price, note, sale_date, sale_time = sale
        formatted_sale_date = datetime.strptime(sale_date, '%Y.%m.%d').strftime('%m.%d')
        profit = selling_price - purchase_price
        row_data = [sale_time, formatted_sale_date, product_name, product_type, product_v1, selling_price, profit, note]
        sheet.append(row_data)

    sheet.cell(row=1, column=10, value='Всего продаж:')
    sheet.cell(row=1, column=11, value=total_sales)
    sheet.cell(row=2, column=10, value='Общая прибыль:')
    sheet.cell(row=2, column=11, value=total_profit)
    sheet.cell(row=3, column=10, value='Дата первой продажи:')
    sheet.cell(row=3, column=11, value=first_sale_date)

    row_offset = 3
    sheet.cell(row=row_offset - 2, column=13, value='По типам продуктов')
    sheet.cell(row=row_offset - 1, column=13, value='Тип')
    sheet.cell(row=row_offset - 1, column=14, value='Количество')
    sheet.cell(row=row_offset - 1, column=15, value='Выручка')

    for i, (product_type, sales_count) in enumerate(type_sales.items(), start=row_offset):
        sheet.cell(row=i, column=13, value=f'{product_type}:')
        sheet.cell(row=i, column=14, value=sales_count)
        sheet.cell(row=i, column=15, value=type_profit[product_type])

    column_widths = [8, 8, 25, 23, 30, 8, 8, 30, 10, 23, 11, 10, 23, 13, 10, 30]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    excel_file_name = 'all_stats.xlsx'
    workbook.save(excel_file_name)

    return excel_file_name

@dp.callback_query(lambda query: query.data == 'all')
async def stat_all(query):
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup = show_menu)
    sql_query = 'SELECT * FROM Stat'
    sql.execute(sql_query)
    sales_data = sql.fetchall()

    total_sales = 0
    total_revenue = 0
    total_profit = 0

    if sales_data:
        first_sale_date = None
        type_sales = defaultdict(int)
        type_revenue = defaultdict(float)
        type_profit = defaultdict(float)

        for sale in sales_data:
            num, product_name, product_type, product_x1, product_x2, product_v1, purchase_price, selling_price, note, sale_date, sale_time = sale

            if first_sale_date is None or sale_date < first_sale_date:
                first_sale_date = sale_date
            response_message = f'Дата первой продажи: {first_sale_date}\n'
            total_sales += 1
            total_revenue += selling_price
            total_revenue = int(total_revenue) if total_revenue.is_integer() else round(total_revenue, 2)

            total_profit += (selling_price - purchase_price)
            total_profit = int(total_profit) if total_profit.is_integer() else round(total_profit, 2)

            type_sales[product_type] += 1

            type_revenue[product_type] += selling_price
            type_revenue[product_type] = int(type_revenue[product_type]) if type_revenue[product_type].is_integer() else round(type_revenue[product_type], 2)

            type_profit[product_type] += (selling_price - purchase_price)
            type_profit[product_type] = int(type_profit[product_type]) if type_profit[product_type].is_integer() else round(type_profit[product_type], 2)

        response_message += f'Всего продаж: {total_sales}\n'
        response_message += f'Выручка (прибыль):\n'
        response_message += f'{total_revenue}р ({total_profit}р)\n'

        response_message += '➖➖➖➖➖➖➖➖➖➖\n'
        for product_type, sales_count in type_sales.items():
            response_message += f'{product_type}\t: {sales_count}\t| {type_revenue[product_type]}р ({type_profit[product_type]}р)\n'

        excel_file = await write_sales_to_excel(sales_data, total_sales, total_profit, first_sale_date, type_sales, type_profit)

    else:
        response_message = 'Нет продаж за все время.'

    await query.message.answer_document(FSInputFile(excel_file), caption=response_message, parse_mode=ParseMode.MARKDOWN)
    await query.message.answer('Стартовое меню', reply_markup=show_menu)

#-----------------------------------------------------------------------------------------------
@dp.callback_query(lambda query: query.data == 'buy')
async def buy(query):
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup = show_menu)
    sql.execute('SELECT c1, buy, sell, type FROM JIJI WHERE ((v1 IS NOT NULL AND c1 > 0))')
    rows = sql.fetchall()

    type_statistics = {}

    total_sales_all = 0
    total_purchase_all = 0
    total_sell_all = 0

    for row in rows:
        c1 = row[0]
        buy = row[1]
        sell = row[2]
        type_ = row[3]

        total_sales_all += c1
        total_purchase_all += c1 * buy
        total_purchase_all = int(total_purchase_all) if total_purchase_all.is_integer() else round(total_purchase_all, 2)
        total_sell_all += c1 * sell
        total_sell_all = int(total_sell_all) if total_sell_all.is_integer() else round(total_sell_all, 2)
        potential_profit_all = total_sell_all - total_purchase_all
        potential_profit_all = int(potential_profit_all) if potential_profit_all.is_integer() else round(potential_profit_all, 2)
        
        response_message = f'Всего товара: {total_sales_all} шт.\n'
        response_message += f'Потрачено: {total_purchase_all}р\n'
        response_message += f'Выручка (прибыль): {total_sell_all}р ({potential_profit_all}р)\n\n'
    
        if type_ not in type_statistics:
            type_statistics[type_] = {'total_sales': 0, 'total_purchase': 0, 'total_sell': 0}

        type_statistics[type_]['total_sales'] += c1
        type_statistics[type_]['total_purchase'] += c1 * buy
        type_statistics[type_]['total_sell'] += c1 * sell

    for type_, stats in type_statistics.items():
        total_sales = stats['total_sales']
        total_purchase = stats['total_purchase']
        total_sell = stats['total_sell']
        potential_profit = total_sell - total_purchase
    
        response_message += f'{type_}: {total_sales} шт.\n'
        response_message += f'Потрачено: {int(total_purchase) if total_purchase.is_integer() else round(total_purchase, 2)}р\n'
        response_message += f'Выручка (прибыль): {int(total_sell) if total_sell.is_integer() else round(total_sell, 2)}р ({int(potential_profit) if potential_profit.is_integer() else round(potential_profit, 2)}р)\n'

    await query.message.edit_text(response_message)
    await query.message.answer('Стартовое меню', reply_markup=show_menu)

#-----------------------------------------------------------------------------------------------
menu_button = [[KeyboardButton(text='Меню')]]
show_menu_button = ReplyKeyboardMarkup(keyboard=menu_button, resize_keyboard=True)
in_menu_button = [[InlineKeyboardButton(text='Меню', callback_data='menu')]]
show_in_menu_button = InlineKeyboardMarkup(inline_keyboard=in_menu_button)

menu_list = [KeyboardButton(text='Меню', resize_keyboard=True)]
in_menu_list = [InlineKeyboardButton(text='Меню', callback_data='menu')]
in_back_sklad = [InlineKeyboardButton(text='Назад', callback_data='sklad'), InlineKeyboardButton(text='МЕНЮ', callback_data='menu')]
#-----------------------------------------------------------------------------------------------
add_status = False

"""smiles = {
    'Никобустер': '❌❌',  # Сложно подобрать смайлики.
    'Мятные леденцы': '🍬🍃',
    'Черная смородина': '🫐🫐',
    'Малиновый лимонад': '🍋🍇',
    'Черничный коктейль': '🫐🥤',
    'Клубника варенье ягоды': '🍓🍓',
    'Яблоко персик': '🍏🍑',
    'Свежесть': '💧💧',
    'Ягоды и кокос': '🍓🥥',
    'Манго и кокос': '🥭🥥',
    'Манго гуава': '🥭🥭',  # Гуава не имеет стандартного смайлика.
    'Дыня банан': '🍈🍌',
    'Голубика лайм': '🫐🍋',
    'Черника жвачка': '🫐🍬',
    'Кола лайм': '🥤🍋',
    'Вишня сгущенка': '🍒🥛',
    'Арбузный милкшейк': '🍉🥤',
    'Грейпфрут лимонад': '🍋🍊',
    'Кокос мороженое': '🥥🍦',
    'Киви и яблоко со льдом': '🥝🍏',
    'Клубника банан маракуйя': '🍓🍌',  # Маракуйя не имеет стандартного смайлика.
    'Яблочная жвачка': '🍏🍬',
    'Вишня с лимоном': '🍒🍋',
    'Голубика малина': '🫐🍇',
    'Банан клубника сливки': '🍌🍓',  # 'Сливки' не имеет стандартного смайлика.
    'Виноград лед': '🍇❄️',
    'Малина личи': '🍇🍇',  # Личи не имеет стандартного смайлика.
    'Сладкая дыня банан': '🍈🍌',
    'Апельсиновый тик так': '🍊🍬', 
    'Маракуйя персик личи': '🥭🍑',  # Личи и маракуйя не имеют стандартного смайлика.
    'Гранат клубника шелковица': '🍇🍓',  # Шелковица не имеет стандартного смайлика.
    'Арбуз киви клубника': '🍉🥝',
    'Груша манго мята': '🍐🥭',
    'Клубничный лимонад': '🍋🍓',
    'Милкшейк груша банан': '🍐🍌',  # Молочный коктейль.
    'Дыня с киви': '🍈🥝',
    'Банан дыня кокос': '🍌🍈',
    'Смородина черника': '🫐🫐',
    'Гранат малина черника': '🍇🍇',  # Гранат и черника.
    'Апельсин свежесть': '🍊💧',
    'Ледяной ананас апельсин': '🍍🍊',
    'Лимонад манго': '🍋🥭',
    'Малина ежевика лед': '🍇❄️',
    'Черника ежевика лед': '🫐❄️',
    'Мятная жвачка': '🍬🍃',
    'Черника малина лимонад': '🫐🍋',
    'Йогурт персик маракуйя': '🍑🥛',
    'Дыня маракуйя': '🍈🍈',  # Маракуйя не имеет стандартного смайлика.
    'Садовые ягоды': '🍓🍓',
    'Арбузный фреш': '🍉🥤',
    'Личи и голубика': '🫐🫐',  # Личи не имеет стандартного смайлика.
    'Виноградный сок': '🍇🍇',
    'Гибрид персик манго': '🥭🍑',
    'Йогурт с ягодами': '🍓🥛',
    'Фрутелла': '🍬🍬',  # Название, указывающее на конфеты.
    'Клубничный мохито': '🍓🍹',
    'Варенье смородина ежевика': '🍇🍇',
    'Голубика смородина': '🫐🍇',
    'Маракуйя мандарин': '🍊🍊',  # Мандарин и маракуйя похожи.
    'Виноград голубика': '🍇🫐',
    'Пина колада': '🍍🥥',
    'Ягодный йогурт': '🍓🥛',
    'Мятный тик так': '🍬🍃',
    'Банан клубника': '🍌🍓',
    'Груша яблоко': '🍐🍏',
    'Груша': '🍐🍐',
    'Дыня': '🍈🍈',
    'Ананас лед': '🍍❄️',
    'Киви лед': '🥝❄️',
    'Банан лед': '🍌❄️',
    'Гранат черешня мята': '🍇🍒',
    'Грейпфрут малина клубника': '🍊🍇',
    'Голубика мята': '🫐🍃',
    'Виноград смородина': '🍇🍇',
    'Дыня банан фрукты': '🍈🍌',
    'Жвачка смородина': '🍬🍇',
    'Арбуз': '🍉🍉',
    'Вишневая содовая': '🍒🥤',
    'Смородина мята': '🍇🍃',
    'Яблоко': '🍏🍏',
    'Арбуз жвачка': '🍉🍬',
    'Клубника арбуз': '🍓🍉',
    'Нектарин вишня': '🍑🍒',
    'Клубничное мороженое': '🍓🍦',
    'Кокос ананас': '🥥🍍',
    'Манго ананас': '🥭🍍',
    'Дыня арбуз': '🍈🍉',
    'Скитлс': '🍬🍬',
    'Вишневый компот': '🍒🥤',
    'Виноград киви': '🍇🥝',
    'Земляника черника': '🍓🫐',
    'Апельсин грейпфрут': '🍊🍊',
    'Малина вишня': '🍇🍒',
    'Алоэ виноград': '🍇🍇',
    'Вишневые леденцы': '🍒🍬',
    'Доктор пеппер апельсин': '🥤🍊',
    'Виноград мята': '🍇🍃',
    'Киви личи': '🥝🥝',  # Личи не имеет стандартного смайлика.
    'Черная вишня': '🍒🍒',
    'Арбуз малина': '🍉🍇',
    'Персик жвачка': '🍑🍬',
    'Персик маракуйя': '🍑🍑',  # Маракуйя не имеет стандартного смайлика.
    'Вишня лайм': '🍒🍋',
    'Дыня банан тропические фрукты': '🍈🍌',
    'Клюквенная газировка': '🍹🍇',  # Клюква заменена.
    'Малина смородина': '🍇🍇',
    'Манго мандарин': '🥭🍊',
    'Ананас яблоко': '🍍🍏',
    'Киви банан': '🥝🍌',
    'Персик манго': '🍑🥭',
    'Красная вишня': '🍒🍒',
    'Смесь африканских фруктов': '🍌🍍',
    'Смородина кислинка': '🍇🍋',
    'Виноградный адреналин раш': '❌❌',  # Адреналин раш имеет авторские права.
    'Вишневая газировка': '🍒🥤',
    'Малиновый йогурт с манго': '🍇🥭',
    'Лимонная газировка с колой': '🍋🥤',
    'Смородиновый коктейль клубника': '🍇🍓',
    'Энергетик ягоды': '🥤🍇',
    'Груша с мятой': '🍐🍃',
    'Мандариновый сорбет': '🍊🍦',
    'Виноградная газировка': '🍇🥤',
    'Голубика дыня': '🫐🍈',
    'Банан дыня кокос': '🍌🍈',
    'Клубника банан': '🍓🍌',
    'Манго кокос': '🥭🥥',
    'Апельсин маракуйя': '🍊🍊',  # Маракуйя заменена.
    'Клюква арония': '🍇🍇',  # Арония заменена.
    'Кактус лайм': '🌵🍋',
    'Клубника конфета': '🍓🍬',
    'Голубика малина': '🫐🍇',
    'Кристальный розовый': '💎🌸',  # Использован символ 'кристалл'.
    'Кристальный голубой': '💎💧',
    'Фиолетово синий': '💜💙',
    'Светло розовый': '🌸🌸',
    'Оранжевый': '🟧🟧',
    'Синий': '🔵🔵',
    'Черный': '⚫⚫',
    'Классический': '⭐⭐',
    'Оригинал': '🔥🔥',
    '0,6 Ом': '❌❌',  # Невозможно подобрать смайлики.
    '0,8 Ом': '❌❌',
    '0,7 Ом': '❌❌',
    '0,2 Ом': '❌❌',
    '0,3 Ом': '❌❌'
}
"""
features = {
            'Жидкости': 'Вкусы:',
            'Расходники': 'Сопротивление:',
            'Многоразовые системы': 'Цвета:',
            'Энергетики': 'Вкусы:',
            'Жевательные смеси': 'Вкусы:'
        }

@dp.callback_query(lambda query: query.data == 'add')
async def add_one(query):
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup=show_menu)
    global add_status
    await query.message.edit_text('Формат:\n1) Раздел (Жидкости)\n2) Название (HOTSPOT Dot)\n3) Характеристика 1 (50mg)\n4) Характеристика 2 (30ml)\n5) Цена покупки\n6) Цена продажи\n7) Вкус\n8) Количество товара', reply_markup=show_in_menu_button)
    add_status = True

async def add_one_message(message):
    if message.from_user.id not in admins:
        return message.answer('Ты не мой братик!', reply_markup=show_menu)
    global add_status
    await message.answer('Формат:\n1) Раздел (Жидкость)\n2) Название (HOTSPOT Dot)\n3) Характеристика 1 (50mg)\n4) Характеристика 2 (30ml)\n5) Цена покупки\n6) Цена продажи\n7) Вкус\n8) Количество товара', reply_markup=show_in_menu_button)
    add_status = True

@dp.message(lambda message: add_status)
async def add(message):
    if message.from_user.id not in admins:
        return await message.answer('Ты не мой братик!', reply_markup = show_menu)
    global add_status
    if add_status == False: return
    lines = message.text.splitlines()
    if len(lines) < 8:
        await message.answer('Неверный формат сообщения.', reply_markup=show_menu)
        return await add_one_message(message)

    type, name, x1, x2, buy, sell, *entries = map(str.strip, lines)

    if len(entries) % 2 != 0:
        await message.answer('Неправильный формат.\nПравильный:\nВкус\nКоличество', reply_markup=show_menu)
        return await add_one_message(message)

    flavors_and_counts = [entries[i:i+2] for i in range(0, len(entries), 2)]
    message_count = 0
    for v1, count in flavors_and_counts:
        count = int(count)
        sql.execute(f'''SELECT name from JIJI WHERE name=(?) AND type=(?) AND v1=(?) AND x1=(?) AND x2=(?) AND buy=(?) AND sell=(?) ''', 
                    (name, type, v1, x1, x2, buy, sell))
        existing_entry = sql.fetchone()
        if existing_entry is None:
            sql.execute(f'''INSERT INTO JIJI (name, type, v1, x1, x2, c1, buy, sell) VALUES (?, ?, ?, ?, ?, ?, ?, ?) ''',
                        (name, type, v1, x1, x2, count, buy, sell))
            db.commit()
            sql.execute(f'''SELECT mess_id from JIJI WHERE name=(?) AND type=(?) AND x1=(?) AND x2=(?) AND sell=(?) ''', 
                (name, type, x1, x2, sell))
            message_id = sql.fetchone()[0]
            if sql.fetchone() is None:
                message_count += 1
                await message.answer(f'Пост {name} создан. ({message_count})')
                await send_new_post(name, type, x1, x2, v1, buy, sell, message_id)
                await zakrep_edit(message)
            else:
                message_count += 1
                await message.answer(f'Пост {name} изменён. ({message_count})')
                await edit_post(name, type, x1, x2, v1, buy, sell, message_id)
        else:
            sql.execute(f'''SELECT mess_id from JIJI WHERE name=(?) AND type=(?) AND v1=(?) AND x1=(?) AND x2=(?) AND buy=(?) AND sell=(?) ''', 
                    (name, type, v1, x1, x2, buy, sell))
            message_id = sql.fetchone()[0]
            message_text = f'Количество {name} {v1} изменено.'
            sql.execute('''SELECT c1 FROM JIJI WHERE name=(?) AND type=(?) AND x1=(?) AND x2=(?) AND v1=(?) AND buy=(?) AND sell=(?) ''', 
                        (name, type, x1, x2, v1, buy, sell))
            c1 = int(sql.fetchone()[0])
            upd_count = c1 + count
            sql.execute(f'''UPDATE JIJI SET c1=(?) WHERE name=(?) AND type=(?) AND x1=(?) AND x2=(?) AND v1=(?) AND buy=(?) AND sell=(?) ''', 
                        (upd_count, name, type, x1, x2, v1, buy, sell))
            db.commit()
            if c1 == 0:
                message_text = f'{name} уже имеется. Вкус добавлен'
                try:
                    await edit_post(name, type, x1, x2, v1, buy, sell, message_id)
                except:
                    await message.answer(f'Пост не требует изменений.')
                try:
                    await zakrep_edit(message)
                except:
                    await message.answer(f'Закреп не требует изменений.')
            await message.answer(message_text)

    add_status = False
    await add_one_message(message)

async def send_new_post(name, type, x1, x2, v1, buy, sell, message_id):
    sql.execute(f'''SELECT name, type, x1, x2, v1, sell, mess_id from JIJI WHERE name=(?) AND type=(?) AND x1=(?) AND x2=(?) AND sell=(?) AND ((v1 IS NOT NULL AND c1 > 0)) ''', 
                (name, type, x1, x2, sell))
    result = sql.fetchall()

    products = {}
    for row in result:
        name = row[0]
        type = row[1]
        x1 = str(row[2])
        x2 = str(row[3])
        v1 = row[4]
        sell = str(row[5])
        message_id = str(row[6])

        product_key = f'{name} {type} {x1} {x2} {sell}'
    
        if product_key not in products:
            products[product_key] = {
                'name': [],
                'type': [],
                'x1': [],
                'x2': [],
                'v1': [],
                'sell': [],
                'mess_id': [],
            }

        if v1:
            products[product_key]['name'].append(name)
            products[product_key]['type'].append(type)
            products[product_key]['x1'].append(x1)
            products[product_key]['x2'].append(x2)
            products[product_key]['v1'].append(v1)
            products[product_key]['sell'].append(sell)
            products[product_key]['mess_id'].append(message_id)

    for product_key, details in products.items():
        name_set = set()
        x1_set = set()
        x2_set = set()
        sell_set = set()
        x1x2_set = set()
        flavors_set = set()
        message_id_set = set()

        for name in details['name']:
            if name[0] not in name_set:
                name_set.add(name[0])

        x1x2 = ''
        for x1 in details['x1']:
            if x1[0] not in x1_set:
                if x1 != '0':
                    x1_set.add(x1[0])
                    x1x2 = (f'{x1}\n')
                    
        for x2 in details['x2']:
            if x2[0] not in x2_set:
                if x2 != '0':
                    x2_set.add(x2[0])
                    x1x2 = (f'{x2}\n')
                    
        if x1 != '0' and x2 != '0':
            if x1x2[0] not in x1x2_set:
                x1x2 = (f'{x1}, {x2}\n')
                x1x2_set.add(x1x2[0])
                x1, x2 = x1x2.split(', ')
                x1_set.add(x1[0])
                x2_set.add(x2[0])
        
        feature = 'Вкусы:'
        for type in details['type']:
            if type in features:  
                feature = features[type]

        flavors = ''
        flavors_p = ''
        for vkus in details['v1']:
            if vkus not in flavors_set:
                sql.execute(f'''SELECT smiles from EMOJI WHERE flavor=(?)  ''', (vkus,))
                if sql.fetchone() is not None:
                    sql.execute(f'''SELECT smiles from EMOJI WHERE flavor=(?)  ''', (vkus,))
                    smiles = str(sql.fetchone()[0])
                    flavors += f'{smiles} {vkus}\n'
                else:
                    flavors += f'{vkus}\n'
                flavors_p += f'{vkus}\n'
                flavors_set.add(vkus)
        flavors_list = flavors_p.split('\n')

        '''flavors = ''
        for vkus in details['v1']:
            flavors += f'{vkus}\n'
            flavors_set.add(vkus)
        flavors_list = flavors.split('\n')'''

        for sell in details['sell']:
            if sell[0] not in sell_set:
                sell_set.add(sell[0])

        for message_id in details['mess_id']:
            if message_id[0] not in message_id_set:
                message_id_set.add(message_id[0])

    text = (f'{name}\n{x1x2}\n{feature}\n{flavors}\nЦена: {sell} рублей')
    try:
        pic = f'pic/{name}.jpg'
        sent_message = await bot.send_photo(channel, FSInputFile(pic), caption=text, disable_notification=True)
        message_id = sent_message.message_id
        for flavor in flavors_list:
            sql.execute('''UPDATE JIJI SET mess_id = ? WHERE name = ? AND type = ? AND v1 = ? AND buy = ? AND sell = ? ''', # AND x1 = ? AND x2 = ?
                    (message_id, name, type, flavor, buy, sell)) # x1, x2,
            db.commit()
    except:
        pic = 'pic/pic.jpg'
        sent_message = await bot.send_photo(channel, FSInputFile(pic), caption=text, disable_notification=True)
        message_id = sent_message.message_id
        for flavor in flavors_list:
            sql.execute('''UPDATE JIJI SET mess_id = ? WHERE name = ? AND type = ? AND v1 = ? AND buy = ? AND sell = ? ''',
                    (message_id, name, type, flavor, buy, sell))
            db.commit()

async def edit_post(name, type, x1, x2, v1, buy, sell, message_id):
    sql.execute(f'''SELECT mess_id from JIJI WHERE name=(?) AND type=(?) AND x1=(?) AND x2=(?) AND sell=(?) ''', 
                (name, type, x1, x2, sell))
    message_id = sql.fetchone()[0]
    sql.execute(f'''SELECT name, type, x1, x2, v1, sell from JIJI WHERE name=(?) AND type=(?) AND x1=(?) AND x2=(?) AND sell=(?) ''', 
            (name, type, x1, x2, sell))
    result = sql.fetchall()
    products = {}
    for row in result:
        name = row[0]
        type = row[1]
        x1 = str(row[2])
        x2 = str(row[3])
        v1 = row[4]
        sell = str(row[5])

        product_key = f'{name} {type} {x1} {x2} {sell}'
    
        if product_key not in products:
            products[product_key] = {
                'name': [],
                'type': [],
                'x1': [],
                'x2': [],
                'v1': [],
                'sell': [],
            }

        if v1:
            products[product_key]['name'].append(name)
            products[product_key]['type'].append(type)
            products[product_key]['x1'].append(x1)
            products[product_key]['x2'].append(x2)
            products[product_key]['v1'].append(v1)
            products[product_key]['sell'].append(sell)

    for product_key, details in products.items():
        name_set = set()
        x1_set = set()
        x2_set = set()
        sell_set = set()
        x1x2_set = set()
        flavors_set = set()
        for name in details['name']:
            if name[0] not in name_set:
                name_set.add(name[0])

        x1x2 = ''
        for x1 in details['x1']:
            if x1[0] not in x1_set:
                if x1 != '0':
                    x1_set.add(x1[0])
                    x1x2 = (f'{x1}\n')
                    
        for x2 in details['x2']:
            if x2[0] not in x2_set:
                if x2 != '0':
                    x2_set.add(x2[0])
                    x1x2 = (f'{x2}\n')
                    
        if x1 != '0' and x2 != '0':
            if x1x2[0] not in x1x2_set:
                x1x2 = (f'{x1}, {x2}\n')
                x1x2_set.add(x1x2[0])
                x1, x2 = x1x2.split(', ')
                x1_set.add(x1[0])
                x2_set.add(x2[0])

        feature = 'Вкусы:'
        for type in details['type']:
            if type in features:  
                feature = features[type]

        flavors = ''
        flavors_p = ''
        for vkus in details['v1']:
            if vkus not in flavors_set:
                sql.execute(f'''SELECT smiles from EMOJI WHERE flavor=(?)  ''', (vkus,))
                if sql.fetchone() is not None:
                    sql.execute(f'''SELECT smiles from EMOJI WHERE flavor=(?)  ''', (vkus,))
                    smiles = str(sql.fetchone()[0])
                    flavors += f'{smiles} {vkus}\n'
                else:
                    flavors += f'{vkus}\n'
                flavors_p += f'{vkus}\n'
                flavors_set.add(vkus)
        flavors_list = flavors_p.split('\n')

        for sell in details['sell']:
            if sell[0] not in sell_set:
                sell_set.add(sell[0])

        text = (f'{name}\n{x1x2}\n{feature}\n{flavors}\nЦена: {sell} рублей')
    try:
        await bot.edit_message_caption(chat_id=channel, message_id=message_id, caption=text)
        for flavor in flavors_list:
            sql.execute('''UPDATE JIJI SET mess_id = ? WHERE name = ? AND type = ? AND v1 = ? AND buy = ? AND sell = ? ''', # AND x1 = ? AND x2 = ?
                    (message_id, name, type, flavor, buy, sell)) # x1, x2,
            db.commit()
    except:
        for flavor in flavors_list:
            sql.execute('''UPDATE JIJI SET mess_id = ? WHERE name = ? AND type = ? AND v1 = ? AND buy = ? AND sell = ? ''', # AND x1 = ? AND x2 = ?
                    (message_id, name, type, flavor, buy, sell)) # x1, x2,
            db.commit()

#-----------------------------------------------------------------------------------------------
del_name_list = []
del_flavor_list = []
del_type_list = []

del_type = False
del_name = False
del_flavor = False
del_price = False

@dp.callback_query(lambda query: query.data == 'delete')
async def delete(query):
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup=show_menu)
    await clear()
    global del_type
    del_type = True

    sql.execute('''SELECT type FROM JIJI WHERE ((v1 IS NOT NULL AND c1 > 0)) ''')
    types = sql.fetchall()
    type_set = set()
    for type in types:
        if type[0] not in type_set:
            del_type_list.append(InlineKeyboardButton(text=type[0], callback_data=type[0]))
            type_set.add(type[0])

    key_list = del_type_list + in_menu_list
    type_keyboard = InlineKeyboardBuilder().row(*key_list, width=3)
    await query.message.edit_text(f'ПРОДАЖА\nВыберите раздел товара', reply_markup=type_keyboard.as_markup())

async def delete_mess(message):
    if message.from_user.id not in admins:
        return await message.answer('Ты не мой братик!', reply_markup = show_menu)
    await clear()
    global del_type
    del_type = True

    sql.execute('''SELECT type FROM JIJI WHERE ((v1 IS NOT NULL AND c1 > 0)) ''')
    types = sql.fetchall()
    type_set = set()
    for type in types:
        if type[0] not in type_set:
            del_type_list.append(InlineKeyboardButton(text=type[0], callback_data=type[0]))
            type_set.add(type[0])

    key_list = del_type_list + in_menu_list
    type_keyboard = InlineKeyboardBuilder().row(*key_list, width=3)
    await bot.send_message(text=f'ПРОДАЖА\nВыберите раздел товара', chat_id=message.from_user.id, reply_markup=type_keyboard.as_markup())

@dp.callback_query(lambda query: query.data in [button.text for button in del_type_list])
async def delete_type(message):
    if message.from_user.id not in admins:
        return await message.answer('Ты не мой братик!', reply_markup = show_menu)
    global del_type, del_name, del_flavor, del_price
    if del_type == False: return
    del_type = False
    del_name = True

    type = message.data
    sql.execute('''DELETE FROM WORK ''')
    db.commit()
    sql.execute('''SELECT name FROM JIJI WHERE type=(?) AND ((v1 IS NOT NULL AND c1 > 0))''', (type,))
    names = sql.fetchall()
    name_set = set()
    for name in names:
        if name[0] not in name_set:
            del_name_list.append(KeyboardButton(text=name[0]))
            name_set.add(name[0])

    key_list = del_name_list + menu_list
    name_keyboard = ReplyKeyboardBuilder().row(*key_list, width=3)
    await bot.edit_message_text(text=f'ПРОДАЖА\nВыберите раздел товара', chat_id=message.from_user.id, message_id=message.message.message_id)
    await bot.send_message(text=f'Раздел: {type}\nВыберите категорию товара', chat_id=message.from_user.id, reply_markup=name_keyboard.as_markup(resize_keyboard=True))

@dp.message(lambda message: message.text in [button.text for button in del_name_list])
async def delete_name(message):
    if message.from_user.id not in admins:
        return await message.answer('Ты не мой братик!', reply_markup = show_menu)
    global del_type, del_name, del_flavor, del_price, del_flavor_list, status, name
    if del_name == False: return
    del_name = False
    del_flavor = True

    sql.execute('''DELETE FROM WORK ''')
    db.commit()
    del_flavor_list = []

    if status == True:
        selected_name = name
    else:
        selected_name = message.text
    sql.execute(f'''SELECT type FROM JIJI WHERE name = '{selected_name}' ''')
    type = sql.fetchall()[0][0]
    sql.execute(f'''SELECT
                    x1, x2, v1, c1, mess_id
                FROM JIJI
                WHERE type = '{type}'
                AND name = '{selected_name}'
                AND ((v1 IS NOT NULL AND c1 > 0)) ''')
    flavors = sql.fetchall()
    sql.execute(f'''SELECT buy FROM JIJI WHERE name = '{selected_name}' ''')
    buy = sql.fetchall()
    sql.execute(f'''SELECT sell FROM JIJI WHERE name = '{selected_name}' ''')
    sell = sql.fetchall()
    x = 0
    for flavor in flavors:
        x1, x2, v, c1, mess_id = flavor[0], flavor[1], flavor[2], flavor[3], flavor[4]
        if v is not None and c1 is not None and c1 > 0:
            button_text = f'{v} ({c1})'
            if type == 'Расходник':
                if x1 != '0':
                    button_text = f'{v} {x1} ({c1})'
                if x2 != '0':
                    button_text = f'{v} {x2} ({c1})'
                if x1 != '0' and x2 !='0':
                    button_text = f'{v} {x1}, {x2} ({c1})'
            callback_text = f'{v}:{buy[x][0]}:{sell[x][0]}'
            del_flavor_list.append(KeyboardButton(text=button_text, callback_data=callback_text))
            sql.execute(f'''INSERT INTO WORK (name, type, x1, x2, v1, c1, buy, sell_db, mess_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?) ''', 
                        (selected_name, type, x1, x2, v, c1, buy[x][0], sell[x][0], mess_id)) 
            db.commit()
        x = x + 1
    key_list = del_flavor_list + menu_list
    flavor_keyboard = ReplyKeyboardBuilder().row(*key_list, width=2)
    await message.answer(f'Раздел: {type}\nКатегория: {selected_name}\nВыберите тип товара', reply_markup=flavor_keyboard.as_markup(resize_keyboard=True))

async def shorten_name(name):
    words = name.split()
    shortened_words = [word[:4].capitalize() for word in words]
    return ' '.join(shortened_words)

@dp.message(lambda message: message.text in [button.text for button in del_flavor_list])
async def delete_flavor(message):
    if message.from_user.id not in admins:
        return await message.answer('Ты не мой братик!', reply_markup = show_menu)
    global del_type, del_name, del_flavor, del_price, del_flavor_list, status, name, message_price
    if del_flavor == False: return
    del_flavor = False
    del_price = True

    sql.execute('''SELECT name FROM WORK ''')
    name = str(sql.fetchall()[0][0])
    sql.execute('''SELECT type FROM WORK ''')
    type = str(sql.fetchall()[0][0])
    sql.execute('''SELECT x1 FROM WORK ''')
    x1 = str(sql.fetchall()[0][0])
    sql.execute('''SELECT x2 FROM WORK ''')
    x2 = str(sql.fetchall()[0][0])
    sql.execute('''SELECT v1 FROM WORK ''')
    flavor = str(sql.fetchall()[0][0])
    sql.execute('''SELECT c1 FROM WORK ''')
    c1 = int(sql.fetchall()[0][0])
    sql.execute('''SELECT buy FROM WORK ''')
    buy = int(sql.fetchall()[0][0])
    sql.execute('''SELECT sell_db FROM WORK ''')
    sell_db = int(sql.fetchall()[0][0])
    for button in del_flavor_list:
        if button.text == message.text:
            select_button = button.callback_data
            selected_flavor, buy, sell = select_button.split(':')

    sql.execute('''DELETE FROM WORK WHERE v1 != ? ''', (selected_flavor,))
    sql.execute('''DELETE FROM WORK WHERE buy != ? ''', (buy,))
    sql.execute('''DELETE FROM WORK WHERE sell_db != ? ''', (sell,))
    db.commit()
    mess_price = await message.answer(f'Продано за (р): {selected_flavor} | {buy}₽')
    message_price = mess_price.message_id

@dp.message(lambda message: re.match(r'^\d+(\.\d+)?( \w+(\s\w+)*)?$', message.text))
async def delete_price(message):
    if message.from_user.id not in admins:
        return await message.answer('Ты не мой братик!', reply_markup = show_menu)
    global del_type, del_name, del_flavor, del_price, del_flavor_list, status, name
    if del_price == False: return
    del_name = True
    del_price = False
    status = True

    note = 0
    text_parts = message.text.split(' ')
    sell = float(text_parts[0])
    if len(text_parts) > 1:
        note = ' '.join(text_parts[1:])

    sql.execute('''SELECT name FROM WORK ''')
    name = str(sql.fetchall()[0][0])
   
    sql.execute('''SELECT type FROM WORK ''')
    type = str(sql.fetchall()[0][0])

    sql.execute('''SELECT x1 FROM WORK ''')
    x1 = str(sql.fetchall()[0][0])

    sql.execute('''SELECT x2 FROM WORK ''')
    x2 = str(sql.fetchall()[0][0])

    sql.execute('''SELECT v1 FROM WORK ''')
    v1 = str(sql.fetchall()[0][0])

    sql.execute('''SELECT c1 FROM WORK ''')
    c1 = int(sql.fetchall()[0][0])

    sql.execute('''SELECT buy FROM WORK ''')
    buy = float(sql.fetchall()[0][0])
    buy = int(buy) if buy.is_integer() else round(buy, 2)

    sql.execute('''SELECT sell_db FROM WORK ''')
    sell_db = float(sql.fetchall()[0][0])
    sell_db = int(sell_db) if sell_db.is_integer() else round(sell_db, 2)

    gain = sell - buy
    gain = int(gain) if gain.is_integer() else round(gain, 2)

    sell = str(sell)
    buy = str(buy)
    gain = str(gain)

    upd_count = c1 - 1
    
    date = datetime.now().strftime('%Y.%m.%d')
    time = datetime.now().strftime('%H:%M')

    sql.execute(f'''UPDATE JIJI SET c1=(?) WHERE name=(?) AND v1=(?) AND buy=(?) AND sell=(?) ''', (upd_count, name, v1, buy, sell_db))
    db.commit()

    sql.execute('''DELETE FROM JIJI WHERE c1 = 0 ''')
    db.commit()

    sql.execute('''SELECT name FROM JIJI WHERE type=(?) AND ((v1 IS NOT NULL AND c1 > 0))''', (type,))
    names = sql.fetchall()

    sql.execute(f'''INSERT INTO Stat (name, type, x1, x2, v1, buy, sell, note, date, time) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?) ''',
                    (name, type, x1, x2, v1, buy, sell, note, date, time))
    db.commit()
    #await message.edit_text(f'Продано:\n{type} - {name}\nВкус: {flavor}\nПрибыль: {gain}р')
    await bot.edit_message_text(text=f'Продано:\n{type} - {name}\nВкус: {v1}\nПрибыль: {gain}р', chat_id=message.from_user.id, message_id=message_price)

    if upd_count == 0:
        sql.execute('''SELECT mess_id FROM WORK ''')
        mess_id = sql.fetchone()[0]
        await message.answer(f'{name} {v1} закончился! <a href="{link}{mess_id}">Пост</a> изменён.')

    sql.execute('''SELECT mess_id FROM WORK ''')
    mess_id = sql.fetchone()[0]
    sql.execute(f'''SELECT name, type, x1, x2, sell from JIJI WHERE mess_id=(?) ''', 
            (mess_id,))
    if sql.fetchone() is None:
        try:
            await bot.delete_message(chat_id=channel, message_id=mess_id)
            sql.execute('''SELECT mess_id FROM WORK ''')
            mess_id = sql.fetchone()[0]
            await message.answer(f'Позиция {name} закончилась! <a href="{link}{mess_id}">Пост</a> удалён.')
        except:
            mess_text = (f'НЕТ В НАЛИЧИИ')
            await bot.edit_message_caption(chat_id=channel, message_id=mess_id, caption=mess_text)
            #await net_v_nalichii(mess_id)
            sql.execute('''SELECT mess_id FROM WORK ''')
            mess_id = sql.fetchone()[0]
            await message.answer(f'Позиция {name} закончилась! Удалите <a href="{link}{mess_id}">пост</a>.')
        try:
            await zakrep_edit(message)
        except:
            await message.answer(f'Закреп не требует изменений.')
        #sql.execute('''SELECT mess_id FROM WORK ''')
        #mess_id = sql.fetchone()[0]
        #await message.answer(f'Позиция {name} закончилась! Удалите <a href="{link}{mess_id}">пост</a>.')
        return await delete_mess(message)
    sql.execute(f'''SELECT name, type, x1, x2, v1, sell from JIJI WHERE mess_id=(?) ''', 
            (mess_id,))
    result = sql.fetchall()
    products = {}
    for row in result:
        name = row[0]
        type = row[1]
        x1 = str(row[2])
        x2 = str(row[3])
        v1 = row[4]
        sell = str(row[5])

        product_key = f'{name} {type} {x1} {x2} {sell}'
    
        if product_key not in products:
            products[product_key] = {
                'name': [],
                'type': [],
                'x1': [],
                'x2': [],
                'v1': [],
                'sell': [],
            }

        if v1:
            products[product_key]['name'].append(name)
            products[product_key]['type'].append(type)
            products[product_key]['x1'].append(x1)
            products[product_key]['x2'].append(x2)
            products[product_key]['v1'].append(v1)
            products[product_key]['sell'].append(sell)
            
    for product_key, details in products.items():
        name_set = set()
        x1_set = set()
        x2_set = set()
        sell_set = set()
        flavors_set = set()
        for name in details['name']:
            if name[0] not in name_set:
                name_set.add(name[0])

        x1x2 = ''
        for x1 in details['x1']:
            if x1[0] not in x1_set:
                if x1 != '0':
                    x1x2 = (f'{x1}\n')
                    
        for x2 in details['x2']:
            if x2[0] not in x2_set:
                if x2 != '0':
                    x1x2 = (f'{x2}\n')
                    
        if x2 != '0' and x1 != '0':
            x1x2 = (f'{x1}, {x2}\n')
            x1, x2 = x1x2.split(', ')
        
        feature = 'Вкусы:'
        for type in details['type']:
            if type in features:  
                feature = features[type]

        flavors = ''
        flavors_p = ''
        for vkus in details['v1']:
            if vkus not in flavors_set:
                sql.execute(f'''SELECT smiles from EMOJI WHERE flavor=(?)  ''', (vkus,))
                if sql.fetchone() is not None:
                    sql.execute(f'''SELECT smiles from EMOJI WHERE flavor=(?)  ''', (vkus,))
                    smiles = str(sql.fetchone()[0])
                    flavors += f'{smiles} {vkus}\n'
                else:
                    flavors += f'{vkus}\n'
                flavors_p += f'{vkus}\n'
                flavors_set.add(vkus)
        flavors_list = flavors_p.split('\n')

        for sell in details['sell']:
            if sell[0] not in sell_set:
                sell_set.add(sell[0])
        mess_text = (f'{name}\n{x1x2}\n{feature}\n{flavors}\nЦена: {sell} рублей')
        try:
            await bot.edit_message_caption(chat_id=channel, message_id=mess_id, caption=mess_text)
        except:
            await message.answer(f'Пост не требует изменений.')
    await delete_mess(message)
    await delete_name(message)

async def net_v_nalichii(mess_id):
    """sql.execute(f'''SELECT name, type, x1, x2, v1, sell from JIJI WHERE mess_id=(?) ''', 
            (mess_id,))
    result = sql.fetchall()
    products = {}
    for row in result:
        name = row[0]
        type = row[1]
        x1 = str(row[2])
        x2 = str(row[3])
        v1 = row[4]
        sell = str(row[5])

        product_key = f'{name} {type} {x1} {x2} {sell}'
    
        if product_key not in products:
            products[product_key] = {
                'name': [],
                'type': [],
                'x1': [],
                'x2': [],
                'v1': [],
                'sell': [],
            }

        if v1:
            products[product_key]['name'].append(name)
            products[product_key]['type'].append(type)
            products[product_key]['x1'].append(x1)
            products[product_key]['x2'].append(x2)
            products[product_key]['v1'].append(v1)
            products[product_key]['sell'].append(sell)
            
    for product_key, details in products.items():
        name_set = set()
        x1_set = set()
        x2_set = set()
        sell_set = set()
        for name in details['name']:
            if name[0] not in name_set:
                name_set.add(name[0])

        x1x2 = ''
        for x1 in details['x1']:
            if x1[0] not in x1_set:
                if x1 != '0':
                    x1x2 = (f'{x1}\n')
                    
        for x2 in details['x2']:
            if x2[0] not in x2_set:
                if x2 != '0':
                    x1x2 = (f'{x2}\n')
                    
        if x2 != '0' and x1 != '0':
            x1x2 = (f'{x1}, {x2}\n')
            x1, x2 = x1x2.split(', ')
        
        feature = 'Вкусы:'
        for type in details['type']:
            if type in features:  
                feature = features[type]

        for sell in details['sell']:
            if sell[0] not in sell_set:
                sell_set.add(sell[0])
        mess_text = (f'{name}\n{x1x2}\n{feature}\nНЕТ В НАЛИЧИИ\n\nЦена: {sell} рублей')
        await bot.edit_message_caption(chat_id=channel, message_id=mess_id, caption=mess_text)"""

#-----------------------------------------------------------------------------------------------
@dp.callback_query(lambda query: query.data == 'sklad')
async def sklad(query):
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup=show_menu)
    await clear()
    sklad_keyboard = [
    [InlineKeyboardButton(text='Статистика', callback_data='stat')],
    [InlineKeyboardButton(text='Наличие', callback_data='availability')],
    [InlineKeyboardButton(text='Группа', callback_data='group')],
    [InlineKeyboardButton(text='База данных', callback_data='database')],
    [InlineKeyboardButton(text='Смайлики', callback_data='smiles')],
    [InlineKeyboardButton(text='Меню', callback_data='menu')]
    ]
    show_sklad_keyboard = InlineKeyboardMarkup(inline_keyboard=sklad_keyboard)
    await query.message.edit_text('Склад', reply_markup=show_sklad_keyboard)

@dp.callback_query(lambda query: query.data == 'group')
async def group(query):
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup = show_menu)
    group_keyboard = [
         [InlineKeyboardButton(text='Посты', callback_data='vse')],
         [InlineKeyboardButton(text='Закреп', callback_data='zakrep')],
         [InlineKeyboardButton(text='Назад', callback_data='sklad'), InlineKeyboardButton(text='Меню', callback_data='menu')]
         ]

    show_group_keyboard = InlineKeyboardMarkup(inline_keyboard=group_keyboard)
    await query.message.edit_text('Склад', reply_markup=show_group_keyboard)

#-----------------------------------------------------------------------------------------------
cancelvse = False

@dp.callback_query(lambda query: query.data == 'vse')
async def vse(query):
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup = show_menu)
    vse_yes_no = [[InlineKeyboardButton(text='Да', callback_data='vse_yes')],
        [InlineKeyboardButton(text='Нет', callback_data='menu')],
        ]
    show_vse_yes_no = InlineKeyboardMarkup(inline_keyboard=vse_yes_no)
    await query.message.edit_text('Запускаем?', reply_markup=show_vse_yes_no)

@dp.callback_query(lambda query: query.data == 'vse_yes')
async def vse_yes(query):
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup=show_menu)
    cancel = [[InlineKeyboardButton(text='Отмена', callback_data='cancelvse')]]
        
    show_cancel = InlineKeyboardMarkup(inline_keyboard=cancel)
    await query.message.edit_text('В процессе...', reply_markup=show_cancel)
    await query.message.answer(f'Что делаем?', reply_markup=show_menu)
    sql.execute('''SELECT name, type, x1, x2, v1, c1, sell FROM JIJI WHERE ((v1 IS NOT NULL AND c1 > 0))''')
    result = sql.fetchall()

    products = {}
    for row in result:
        name = row[0]
        type = row[1]
        x1 = str(row[2])
        x2 = str(row[3])
        v1 = row[4]
        c1 = row[5]
        sell = str(row[6])

        product_key = f'{name} {type} {x1} {x2} {sell}'
    
        if product_key not in products:
            products[product_key] = {
                'name': [],
                'type': [],
                'x1': [],
                'x2': [],
                'v1': [],
                'c1': [],
                'sell': [],
            }

        if v1:
            products[product_key]['name'].append(name)
            products[product_key]['type'].append(type)
            products[product_key]['x1'].append(x1)
            products[product_key]['x2'].append(x2)
            products[product_key]['v1'].append(v1)
            products[product_key]['c1'].append(c1)
            products[product_key]['sell'].append(sell)

    for product_key, details in products.items():
        name_set = set()
        x1_set = set()
        x2_set = set()
        sell_set = set()
        x1x2_set = set()
        vkus_set = set()
        flavors_set = set()
        for name in details['name']:
            if name[0] not in name_set:
                name_set.add(name[0])

        x1x2 = ''
        for x1 in details['x1']:
            if x1[0] not in x1_set:
                if x1 != '0':
                    x1_set.add(x1[0])
                    x1x2 = (f'{x1}\n')
                    
        for x2 in details['x2']:
            if x2[0] not in x2_set:
                if x2 != '0':
                    x2_set.add(x2[0])
                    x1x2 = (f'{x2}\n')
                    
        if x1 != '0' and x2 != '0':
            if x1x2[0] not in x1x2_set:
                x1x2 = (f'{x1}, {x2}\n')
                x1x2_set.add(x1x2[0])
                x1, x2 = x1x2.split(', ')
                x1_set.add(x1[0])
                x2_set.add(x2[0])

        feature = 'Вкусы:'  
        for type in details['type']:
            if type in features:  
                feature = features[type]

        '''flavors = ''
        vkuskolvo = ''
        for vkus, kolvo in zip(details['v1'], details['c1']):
            if vkus not in vkus_set:
                vkuskolvo += f'{vkus}\n'
                vkus_set.add(vkus)
                flavors = '\n'.join(details['v1'])
        flavors_list = flavors.split('\n')'''

        flavors = ''
        flavors_p = ''
        for vkus in details['v1']:
            if vkus not in flavors_set:
                sql.execute(f'''SELECT smiles from EMOJI WHERE flavor=(?)  ''', (vkus,))
                if sql.fetchone() is not None:
                    sql.execute(f'''SELECT smiles from EMOJI WHERE flavor=(?)  ''', (vkus,))
                    smiles = str(sql.fetchone()[0])
                    flavors += f'{smiles} {vkus}\n'
                else:
                    flavors += f'{vkus}\n'
                flavors_p += f'{vkus}\n'
                flavors_set.add(vkus)
        flavors_list = flavors_p.split('\n')

        for sell in details['sell']:
            if sell[0] not in sell_set:
                sell_set.add(sell[0])
        text = (f'{name}\n{x1x2}\n{feature}\n{flavors}\nЦена: {sell} рублей')
        try:
            pic = f'pic/{name}.jpg'
            sent_message = await bot.send_photo(channel, FSInputFile(pic), caption=text, disable_notification=True)
            message_id = sent_message.message_id
            for flavor in flavors_list:
                sql.execute('''UPDATE JIJI SET mess_id = ? WHERE name = ? AND type = ? AND v1 = ? AND sell = ? ''', # AND x1 = ? AND x2 = ?
                        (message_id, name, type, flavor, sell)) # x1, x2,
                db.commit()
        except:
            pic = 'pic/pic.jpg'
            sent_message = await bot.send_photo(channel, FSInputFile(pic), caption=text, disable_notification=True)
            message_id = sent_message.message_id
            for flavor in flavors_list:
                sql.execute('''UPDATE JIJI SET mess_id = ? WHERE name = ? AND type = ? AND v1 = ? AND sell = ? ''',
                        (message_id, name, type, flavor, sell))
                db.commit()
        global cancelvse
        if cancelvse == True:
            cancelvse = False
            return await query.message.edit_text('Скрипт остановлен')
        await asyncio.sleep(4)

    await query.message.edit_text('Готово!')
    await zakrep(query)

#-----------------------------------------------------------------------------------------------
@dp.callback_query(lambda query: query.data == 'availability')
async def availability(query):
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup = show_menu)
    sql.execute('''SELECT name, type, x1, x2, v1, c1, sell FROM JIJI WHERE ((v1 IS NOT NULL AND c1 > 0))''')
    result = sql.fetchall()

    products = {}
    for row in result:
        name = row[0]
        type = row[1]
        x1 = str(row[2])
        x2 = str(row[3])
        v1 = row[4]
        c1 = row[5]
        sell = str(row[6])

        product_key = f'{name} {type} {x1} {x2} {sell}'
    
        if product_key not in products:
            products[product_key] = {
                'name': [],
                'type': [],
                'x1': [],
                'x2': [],
                'v1': [],
                'c1': [],
                'sell': [],
            }

        if v1:
            products[product_key]['name'].append(name)
            products[product_key]['type'].append(type)
            products[product_key]['x1'].append(x1)
            products[product_key]['x2'].append(x2)
            products[product_key]['v1'].append(v1)
            products[product_key]['c1'].append(c1)
            products[product_key]['sell'].append(sell)

    messagetext = ''
    for product_key, details in products.items():
        name_set = set()
        sell_set = set()
        x2_set = set()
        x1_set = set()
        for name in details['name']:
            if name[0] not in name_set:
                messagetext += f'\n{name}\n'
                name_set.add(name[0])
        text = ''
        for x1 in details['x1']:
            if x1[0] not in x1_set:
                if x1 != '0':
                    text = (f'{x1}\n')
                    x1_set.add(x1[0]) 
        for x2 in details['x2']:
            if x2[0] not in x2_set:
                if x2 != '0':
                    text = (f'{x2}\n')
                    x2_set.add(x2[0])
        if x2 != '0' and x1 != '0':
            text = (f'{x1}, {x2}\n')
        messagetext += f'{text}'
        
        feature = 'Вкусы:'
        for type in details['type']:
            if type in features:  
                feature = features[type]

        messagetext += f'{feature}\n'   
        for vkus, kolvo in zip(details['v1'], details['c1']):
            messagetext += f'{vkus}: {kolvo}\n'
        for sell in details['sell']:
            if sell[0] not in sell_set:
                messagetext += f'Цена: {sell} рублей\n'
                sell_set.add(sell[0])
    await query.message.edit_text(messagetext, reply_markup = show_menu) # f'{name}\n{text}\n{feature}:\n{vkus}: {kolvo}'

#-----------------------------------------------------------------------------------------------
async def zakrep_edit(message):
    if message.from_user.id not in admins:
        return await message.answer('Ты не мой братик!', reply_markup=show_menu)
    
    sql.execute('''SELECT DISTINCT x1 FROM JIJI WHERE type = 'Жидкости' ORDER BY x1 DESC, name ASC ''')
    x1 = sql.fetchall()

    message = '<b><i>АКТУАЛЬНОЕ НАЛИЧИЕ:</i></b>\n'
    liquids = {}

    for mg_value in x1:
        mg = mg_value[0]
        sql.execute(f'''SELECT name, mess_id, sell FROM JIJI WHERE x1 = ? AND type = 'Жидкости' ORDER BY name ASC ''', (mg,))
        result = sql.fetchall()
        if mg in ['45mg', '50mg', '60mg']:
            mg = '~50mg'
        if mg not in liquids:
            liquids[mg] = []
        if result:
            for name, mess_id, sell in result:
                if (name, mess_id, sell) not in liquids[mg]:
                    liquids[mg].append((name, mess_id, sell))

    for mg, liquid_list in liquids.items():
        message += f'\n<u>— Жидкости {mg}</u>\n'
        for name, mess_id, sell in liquid_list:
            message += f'<u><a href="{link}{mess_id}">{name}</a></u> {sell}р\n'

    sql.execute('''SELECT DISTINCT type FROM JIJI WHERE type != 'Жидкости' ORDER BY name ASC ''')
    other_types = sql.fetchall()

    for type_value in other_types:
        other_type = type_value[0]
        message += f'\n<u>— {other_type}</u>\n'
        unique_entries = []
        if other_type == 'Расходники':
            sql.execute('''SELECT name, mess_id, v1, sell FROM JIJI WHERE type = ? ORDER BY name ASC ''', (other_type,))
            result = sql.fetchall()
            if result:
                for name, mess_id, v1, sell in result:
                    entry = (name, v1)
                    if entry not in unique_entries:
                        message += f'<u><a href="{link}{mess_id}">{name} {v1}</a></u> {sell}р\n'
                        unique_entries.append(entry)
        else:
            sql.execute(f'''SELECT name, mess_id, sell FROM JIJI WHERE type = ? ORDER BY name ASC ''', (other_type,))
            result = sql.fetchall()
            if result:
                for name, mess_id, sell in result:
                    if name not in unique_entries:
                        message += f'<u><a href="{link}{mess_id}">{name}</a></u> {sell}р\n'
                        unique_entries.append(name)

    message += '\nКонтакт для связи — @tosh1nn —'
    sql.execute(f'''SELECT zakrep_id FROM Zakrep ''')
    id = sql.fetchone()[0]
    await bot.edit_message_text(text=message, chat_id=channel, message_id=id, parse_mode='HTML')

@dp.callback_query(lambda query: query.data == 'zakrep')
async def zakrep(query):
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup=show_menu)
    
    sql.execute('''SELECT DISTINCT x1 FROM JIJI WHERE type = 'Жидкости' ORDER BY x1 DESC, name ASC ''')
    x1 = sql.fetchall()

    message = '<b><i>АКТУАЛЬНОЕ НАЛИЧИЕ:</i></b>\n'
    liquids = {}

    for mg_value in x1:
        mg = mg_value[0]
        sql.execute(f'''SELECT name, mess_id, sell FROM JIJI WHERE x1 = ? AND type = 'Жидкости' ORDER BY name ASC ''', (mg,))
        result = sql.fetchall()
        if mg in ['45mg', '50mg', '60mg']:
            mg = '~50mg'
        if mg not in liquids:
            liquids[mg] = []
        if result:
            for name, mess_id, sell in result:
                if (name, mess_id, sell) not in liquids[mg]:
                    liquids[mg].append((name, mess_id, sell))

    for mg, liquid_list in liquids.items():
        message += f'\n<u>— Жидкости {mg}</u>\n'
        for name, mess_id, sell in liquid_list:
            message += f'<u><a href="{link}{mess_id}">{name}</a></u> {sell}р\n'

    sql.execute('''SELECT DISTINCT type FROM JIJI WHERE type != 'Жидкости' ORDER BY name ASC ''')
    other_types = sql.fetchall()

    for type_value in other_types:
        other_type = type_value[0]
        message += f'\n<u>— {other_type}</u>\n'
        unique_entries = []
        if other_type == 'Расходники':
            sql.execute('''SELECT name, mess_id, v1, sell FROM JIJI WHERE type = ? ORDER BY name ASC ''', (other_type,))
            result = sql.fetchall()
            if result:
                for name, mess_id, v1, sell in result:
                    entry = (name, v1)
                    if entry not in unique_entries:
                        message += f'<u><a href="{link}{mess_id}">{name} {v1}</a></u> {sell}р\n'
                        unique_entries.append(entry)
        else:
            sql.execute(f'''SELECT name, mess_id, sell FROM JIJI WHERE type = ? ORDER BY name ASC ''', (other_type,))
            result = sql.fetchall()
            if result:
                for name, mess_id, sell in result:
                    if name not in unique_entries:
                        message += f'<u><a href="{link}{mess_id}">{name}</a></u> {sell}р\n'
                        unique_entries.append(name)

    message += '\nКонтакт для связи — @tosh1nn —'
    await query.message.edit_text('Сделано!', reply_markup=show_menu)
    try:
        sql.execute('''SELECT zakrep_id FROM Zakrep''')
        id = sql.fetchone()[0]
        await bot.unpin_chat_message(channel, id)
    except:
        await query.message.edit_text(f'Не удалось открепить.', reply_markup=show_menu)
    zakrep_id = await bot.send_message(channel, message, parse_mode='HTML')
    message_id = zakrep_id.message_id
    await bot.pin_chat_message(channel, message_id)
    sql.execute('''UPDATE Zakrep SET zakrep_id=(?)''', 
                    (message_id,))
    db.commit()

#-----------------------------------------------------------------------------------------------
@dp.callback_query(lambda query: query.data == 'cancelvse')
async def canselvse(query):
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup = show_menu)
    global cancelvse
    cancelvse = True

"""@dp.callback_query(lambda query: query.data == 'cancelLast')
async def cancellast(query):
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup = show_menu)
    sql.execute('''SELECT name, type, x1, x2, v1, buy, sell FROM Stat ORDER BY ROWID DESC LIMIT 1 ''')
    result = sql.fetchall()
    for row in result:
        name = row[0]
        type = row[1]
        x2 = str(row[2])
        x1 = str(row[3])
        v1 = row[4]
        sell = str(row[5])
    try:
        can_last = [[InlineKeyboardButton(text='Да', callback_data='cancelLastYes')],
        [InlineKeyboardButton(text='Нет', callback_data='menu')],
        ]
        show_can_last = InlineKeyboardMarkup(inline_keyboard=can_last)
        await query.message.edit_text(f'Удалить последнее?\n{name}\n{type}\n{v1}\n{sell}₽', reply_markup=show_can_last)
    except:
        await query.message.edit_text(f'Записей нет', reply_markup=show_menu)

@dp.callback_query(lambda query: query.data == 'cancelLastYes')
async def cancellastyes(query):
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup = show_menu)
    sql.execute('''SELECT name, type, x1, x2, v1, buy, sell FROM Stat ORDER BY ROWID DESC LIMIT 1 ''')
    result = sql.fetchall()
    for row in result:
        name = row[0]
        type = row[1]
        x2 = str(row[2])
        x1 = str(row[3])
        v1 = row[4]
        buy = str(row[5])
        sell = str(row[6])
    try:
        sql.execute('''DELETE FROM Stat WHERE ROWID = (SELECT MAX(ROWID) FROM Stat) ''')
        db.commit()
        sql.execute(f'''SELECT c1 from JIJI WHERE name=(?) AND type=(?) AND v1=(?) AND x1=(?) AND x2=(?) AND buy=(?) AND sell=(?) ''', 
                    (name, type, v1, x1, x2, buy, sell))
        if sql.fetchone() is None:
            sql.execute(f'''INSERT INTO JIJI (name, type, v1, x1, x2, c1, buy, sell) VALUES (?, ?, ?, ?, ?, ?, ?, ?) ''',
                        (name, type, v1, x1, x2, 1, buy, sell))
            db.commit()
        else:
            sql.execute(f'''SELECT c1 from JIJI WHERE name=(?) AND type=(?) AND v1=(?) AND x1=(?) AND x2=(?) AND buy=(?) AND sell=(?) ''', 
                        (name, type, v1, x1, x2, buy, sell))
            count = int(sql.fetchone()[0])
            upd_count = count + 1
            sql.execute(f'''UPDATE JIJI SET c1=(?) WHERE name=(?) AND type=(?) AND x1=(?) AND x2=(?) AND v1=(?) AND buy=(?) AND sell=(?) ''', 
                            (upd_count, name, type, x1, x2, v1, buy, sell))
            db.commit()
        await query.message.edit_text(f'Последняя продажа отменена\n{name}\n{type}\n{v1}\n{sell}₽', reply_markup=show_menu)
    except:
        await query.message.edit_text(f'Продаж нет', reply_markup=show_menu)
"""
"""@dp.callback_query(lambda query: query.data == 'cancelLast')
async def cancellast(query):
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup=show_menu)

    # Получаем последние 10 записей из таблицы `Stat` по столбцу `num`
    sql.execute('''
        SELECT num, name, type, x1, x2, v1, buy, sell 
        FROM Stat 
        ORDER BY num DESC 
        LIMIT 10
    ''')
    records = sql.fetchall()

    if not records:
        return await query.message.answer("В базе данных нет записей.")

    # Формируем сообщение
    message = "\n".join([
        f"{record[0]}) {record[1]}, {record[2]}, {record[3]}, {record[4]}, {record[5]}, {record[6]}, {record[7]}"
        for record in records
    ])

    message += "\n\nВведите номер `num`, чтобы удалить запись из базы данных."

    await query.message.answer(message)

@dp.message(lambda message: message.text.isdigit())
async def delete_record(message):
    num = int(message.text)

    sql.execute("SELECT * FROM Stat WHERE num = ?", (num,))
    record = sql.fetchone()

    if not record:
        return await message.answer("Запись с указанным номером не найдена.")
    sql.execute("DELETE FROM Stat WHERE num = ?", (num,))
    db.commit()

    await message.answer(f"Запись с номером {num} успешно удалена.")
"""

del_last_list = []
@dp.callback_query(lambda query: query.data == 'cancelLast')
async def cancellast(query):
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup=show_menu)

    sql.execute('''
        SELECT num, name, type, x1, x2, v1, buy, sell 
        FROM Stat 
        ORDER BY num DESC 
        LIMIT 10
    ''')
    records = sql.fetchall()

    if not records:
        return await query.message.answer("В базе данных нет записей.")

    for record in records:
        button_text = f"{record[0]}) {record[1]}, {record[2]}, {record[3]}, {record[4]}, {record[5]}, {record[6]}, {record[7]}"
        callback_data = f"delete_{record[0]}"
        del_last_list.append(InlineKeyboardButton(text=button_text, callback_data=callback_data))
    key_list = del_last_list + in_back_sklad
    del_last_keyboard = InlineKeyboardBuilder().row(*key_list, width=1)
    
    await query.message.edit_text("Выберите запись для удаления:", reply_markup=del_last_keyboard.as_markup())

@dp.callback_query(lambda query: query.data.startswith("delete_"))
async def confirm_delete(query):
    num = int(query.data.split("_")[1])

    sql.execute("SELECT name, type, v1, buy, sell FROM Stat WHERE num = ?", (num,))
    record = sql.fetchone()

    if not record:
        return await query.message.answer("Запись с указанным номером не найдена.")

    name, record_type, v1, buy, sell = record
    message_text = (
        f"Вы уверены, что хотите удалить запись?\n\n"
        f"Название: {name}\n"
        f"Тип: {record_type}\n"
        f"V1: {v1}\n"
        f"Покупка: {buy}\n"
        f"Продажа: {sell}"
    )

    confirm_keyboard = [[InlineKeyboardButton(text='Да', callback_data=f'confirm_yes_{num}')],
         [InlineKeyboardButton(text='Нет', callback_data='confirm_no')],
         ]

    show_confirm_keyboard = InlineKeyboardMarkup(inline_keyboard=confirm_keyboard)

    await query.message.edit_text(message_text, reply_markup=show_confirm_keyboard)


@dp.callback_query(lambda query: query.data.startswith("confirm_yes_"))
async def delete_record(query):
    num = int(query.data.split("_")[2])

    sql.execute("SELECT name, type, v1, buy, sell FROM Stat WHERE num = ?", (num,))
    record = sql.fetchone()
    name, record_type, v1, buy, sell = record
    message_text = (
        f"Удалено\n\n"
        f"Название: {name}\n"
        f"Тип: {record_type}\n"
        f"V1: {v1}\n"
        f"Покупка: {buy}\n"
        f"Продажа: {sell}"
    )
    sql.execute("DELETE FROM Stat WHERE num = ?", (num,))
    db.commit()

    await query.message.edit_text(message_text, reply_markup=show_menu)
    await clear()

@dp.callback_query(lambda query: query.data == "confirm_no")
async def cancel_delete(query):
    await query.message.edit_text("Удаление отменено", reply_markup=show_menu)
    await clear()
    
@dp.callback_query(lambda query: query.data == 'menu')
async def back_in_menu_query(query):
    if query.from_user.id not in admins:
        return await query.message.answer('Ты не мой братик!', reply_markup=show_menu)
    await clear()
    await connect()
    await query.message.edit_text(f'Стартовое меню', reply_markup=show_menu)

@dp.message(F.text =='Меню')
async def back_in_menu_message(message):
    if message.from_user.id not in admins:
        return await message.answer('Ты не мой братик!', reply_markup=show_menu)
    await clear()
    await connect()
    await message.answer(f'Что делаем?', reply_markup=show_menu)

@dp.callback_query(lambda query: query.data == 'smiles')
async def smiles(query):
    message = ''
    flavor_set = set()
    sql.execute('''SELECT v1 
    FROM JIJI
    WHERE v1 NOT IN (SELECT flavor FROM EMOJI) ''')
    result = sql.fetchall()
    for row in result:
        if row[0] not in flavor_set:
            message += f'{row[0]}\n'
            flavor_set.add(row[0])
        #print(row[0])
    await query.message.edit_text(message, reply_markup=show_menu)
    
@dp.message(Command('start', 's'))
async def send_welcome(message):
    if message.from_user.id not in admins:
        return await message.answer('Ты не мой братик!', reply_markup=show_menu)
    await clear()
    await connect()
    await message.answer(f'<b><i>Что делаем?</i></b>', parse_mode='HTML', reply_markup = show_menu, disable_web_page_preview=True)

async def clear():
    global add_status
    global del_name_list, del_flavor_list, del_type_list
    global del_type, del_name, del_flavor, del_price
    global status
    global send
    global del_last_list

    send = 0

    del_name_list = []
    del_flavor_list = []
    del_type_list = []

    add_status = False

    del_type = False
    del_name = False
    del_flavor = False
    del_price = False

    del_last_list = []
    try:
        sql.execute('''DELETE FROM WORK ''')
        db.commit()
    except:
        pass

    status = False

def lv():
    db = sqlite3.connect('Jiji TEST.db')
    sql = db.cursor()
    sql.execute("SELECT rowid FROM Stat ORDER BY rowid")
    rows = sql.fetchall()

    for index, (rowid,) in enumerate(rows, start=1):
        sql.execute(f'''UPDATE Stat SET num=(?) WHERE rowid=(?) ''', (index, rowid))
    db.commit()

async def main():
    global bot
    await connect()
    bot = Bot(TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
    await dp.start_polling(bot)
    
if __name__ == '__main__':
    asyncio.run(main())
